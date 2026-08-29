from datetime import date, datetime, timedelta
from io import BytesIO
import json
import re

from fastapi import APIRouter, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, or_, select

from bot.db import get_async_session_factory
from bot.models import Booking, Client, DailyPayment, DeletedItem, Item, Preorder, Purchase, Sale
from web_admin.templates import templates

router = APIRouter()

# Максимальная длина произвольного периода (защита от слишком тяжёлого графика)
MAX_RANGE_DAYS = 366


def _normalize_source(raw: str | None) -> str:
    if not raw:
        return "Не указан"
    s = re.sub(r"\s+", " ", str(raw).strip())
    if not s:
        return "Не указан"
    low = s.lower()

    mapping = [
        (("авито", "avito"), "Авито"),
        (("telegram", "телеграм", "тг", "tg"), "Telegram"),
        (("whatsapp", "ватсап", "вацап", "wa"), "WhatsApp"),
        (("знакомые", "посоветовали", "рекоменд", "сарафан"), "Рекомендации"),
        (("уже покупали", "повтор", "постоянн"), "Уже покупали"),
        (("instagram", "инстаграм", "инста"), "Instagram"),
        (("youtube", "ютуб"), "YouTube"),
        (("сайт", "site", "web"), "Сайт"),
        (("офлайн", "магазин", "пришёл", "пришел"), "Офлайн / магазин"),
    ]
    for keys, label in mapping:
        if any(k in low for k in keys):
            return label
    return s[:40] if len(s) > 40 else s


def _safe_parse_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(str(value).strip()[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _parse_period(
    target_date: str | None,
    days: int,
    mode: str,
    month: str | None,
    date_from: str | None,
    date_to: str | None,
) -> tuple[date, date, str]:
    """
    Режимы:
      - preset: последние N дней включая сегодня (target_date или today)
      - month: календарный месяц YYYY-MM
      - range / произвольные date_from + date_to
    Если date_from > date_to — меняем местами.
    Период режется до MAX_RANGE_DAYS.
    """
    today = date.today()
    target = _safe_parse_date(target_date) or today
    mode = (mode or "preset").strip().lower()

    try:
        days = int(days)
    except (TypeError, ValueError):
        days = 7
    days = max(1, min(days, MAX_RANGE_DAYS))

    start_date: date
    end_date: date

    if mode == "preset":
        end_date = target
        start_date = end_date - timedelta(days=days - 1)
    elif mode == "month" and month:
        try:
            year, mon = map(int, month.split("-"))
            start_date = date(year, mon, 1)
            if mon == 12:
                end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(year, mon + 1, 1) - timedelta(days=1)
        except (ValueError, TypeError):
            end_date = today
            start_date = end_date - timedelta(days=6)
    else:
        # range / любое другое: свободные даты
        df = _safe_parse_date(date_from)
        dt = _safe_parse_date(date_to)
        if df and dt:
            start_date, end_date = df, dt
        elif df and not dt:
            start_date = end_date = df
        elif dt and not df:
            start_date = end_date = dt
        else:
            # fallback — 7 дней
            end_date = today
            start_date = end_date - timedelta(days=6)

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    # ограничение длины
    span = (end_date - start_date).days + 1
    if span > MAX_RANGE_DAYS:
        start_date = end_date - timedelta(days=MAX_RANGE_DAYS - 1)

    period_label = f"{start_date.strftime('%d.%m.%Y')} — {end_date.strftime('%d.%m.%Y')}"
    return start_date, end_date, period_label


async def _collect_report(session, start_date: date, end_date: date) -> dict:
    sales_data = await session.execute(
        select(
            func.coalesce(func.sum(Sale.cash), 0).label("cash"),
            func.coalesce(func.sum(Sale.terminal), 0).label("terminal"),
            func.coalesce(func.sum(Sale.qr), 0).label("qr"),
            func.coalesce(func.sum(Sale.transfer), 0).label("transfer"),
            func.coalesce(func.sum(Sale.invoice), 0).label("invoice"),
            func.coalesce(func.sum(Sale.installment), 0).label("installment"),
            func.count(Sale.id).label("count"),
        ).where(func.date(Sale.sold_at).between(start_date, end_date))
    )
    sales_row = sales_data.mappings().one()

    preorders_data = await session.execute(
        select(
            func.coalesce(func.sum(Preorder.cash), 0).label("cash"),
            func.coalesce(func.sum(Preorder.terminal), 0).label("terminal"),
            func.coalesce(func.sum(Preorder.qr), 0).label("qr"),
            func.coalesce(func.sum(Preorder.transfer), 0).label("transfer"),
            func.coalesce(func.sum(Preorder.invoice), 0).label("invoice"),
            func.coalesce(func.sum(Preorder.installment), 0).label("installment"),
            func.count(Preorder.id).label("count"),
        ).where(func.date(Preorder.created_at).between(start_date, end_date))
    )
    preorders_row = preorders_data.mappings().one()

    bookings_data = await session.execute(
        select(
            func.coalesce(func.sum(Booking.total_amount), 0).label("total"),
            func.count(Booking.id).label("count"),
        ).where(func.date(Booking.booked_at).between(start_date, end_date))
    )
    bookings_row = bookings_data.mappings().one()

    # График по дням — один запрос вместо N×2
    sales_by_day = {
        row.d: int(row.cnt)
        for row in (
            await session.execute(
                select(
                    func.date(Sale.sold_at).label("d"),
                    func.count(Sale.id).label("cnt"),
                )
                .where(func.date(Sale.sold_at).between(start_date, end_date))
                .group_by(func.date(Sale.sold_at))
            )
        ).all()
    }
    revenue_by_day = {
        row.d: float(row.amt or 0)
        for row in (
            await session.execute(
                select(
                    func.date(DailyPayment.created_at).label("d"),
                    func.coalesce(func.sum(DailyPayment.amount), 0).label("amt"),
                )
                .where(func.date(DailyPayment.created_at).between(start_date, end_date))
                .group_by(func.date(DailyPayment.created_at))
            )
        ).all()
    }

    chart_dates: list[str] = []
    chart_sales: list[int] = []
    chart_revenue: list[float] = []
    current = start_date
    while current <= end_date:
        chart_dates.append(current.strftime("%d.%m"))
        # key may be date or datetime.date from DB driver
        chart_sales.append(int(sales_by_day.get(current, 0) or 0))
        chart_revenue.append(float(revenue_by_day.get(current, 0) or 0))
        current += timedelta(days=1)

    purchase_rows = (
        await session.execute(
            select(
                Client.referral_source,
                Client.social_network,
                Purchase.total_amount,
                Purchase.purchase_type,
                Purchase.items_json,
            )
            .select_from(Purchase)
            .outerjoin(Client, Client.id == Purchase.client_id)
            .where(func.date(Purchase.created_at).between(start_date, end_date))
        )
    ).all()

    source_agg: dict[str, dict] = {}
    model_agg: dict[str, dict] = {}

    for ref, social, amount, ptype, items_json in purchase_rows:
        src = _normalize_source(ref or social)
        if src not in source_agg:
            source_agg[src] = {"count": 0, "amount": 0.0}
        source_agg[src]["count"] += 1
        source_agg[src]["amount"] += float(amount or 0)

        items = []
        if items_json:
            try:
                raw = json.loads(items_json) if isinstance(items_json, str) else items_json
                if isinstance(raw, list):
                    items = raw
                elif isinstance(raw, dict):
                    items = [raw]
            except Exception:
                items = []
        for it in items:
            name = (it.get("item_text") or it.get("text") or it.get("name") or "").strip()
            if not name:
                continue
            if name.lower().startswith("trade-in"):
                continue
            price = float(it.get("price") or 0)
            if name not in model_agg:
                model_agg[name] = {"count": 0, "amount": 0.0}
            model_agg[name]["count"] += 1
            model_agg[name]["amount"] += price

    if len(model_agg) < 5:
        deleted_rows = (
            await session.execute(
                select(DeletedItem.text, func.count(DeletedItem.id))
                .where(
                    func.date(DeletedItem.deleted_at).between(start_date, end_date),
                    or_(
                        DeletedItem.reason.ilike("%sale%"),
                        DeletedItem.sale_message_id.isnot(None),
                    ),
                )
                .group_by(DeletedItem.text)
                .order_by(func.count(DeletedItem.id).desc())
                .limit(30)
            )
        ).all()
        for text, cnt in deleted_rows:
            name = (text or "").strip()
            if not name:
                continue
            if name not in model_agg:
                model_agg[name] = {"count": 0, "amount": 0.0}
            if model_agg[name]["count"] == 0:
                model_agg[name]["count"] = int(cnt or 0)

    booking_platform_rows = (
        await session.execute(
            select(Item.booking_platform, func.count(Item.id))
            .where(Item.is_booked.is_(True))
            .group_by(Item.booking_platform)
        )
    ).all()
    booking_sources: dict[str, int] = {}
    for plat, cnt in booking_platform_rows:
        src = _normalize_source(plat)
        booking_sources[src] = booking_sources.get(src, 0) + int(cnt or 0)

    sources_sorted = sorted(source_agg.items(), key=lambda x: x[1]["amount"], reverse=True)
    models_sorted = sorted(model_agg.items(), key=lambda x: x[1]["count"], reverse=True)[:20]

    return {
        "sales_row": dict(sales_row),
        "preorders_row": dict(preorders_row),
        "bookings_row": dict(bookings_row),
        "chart_dates": chart_dates,
        "chart_sales": chart_sales,
        "chart_revenue": chart_revenue,
        "source_labels": [k for k, _ in sources_sorted],
        "source_counts": [v["count"] for _, v in sources_sorted],
        "source_amounts": [round(v["amount"], 0) for _, v in sources_sorted],
        "sources_table": [
            {"name": n, "count": d["count"], "amount": d["amount"]}
            for n, d in sources_sorted
        ],
        "booking_sources_table": sorted(
            [{"name": k, "count": v} for k, v in booking_sources.items()],
            key=lambda x: x["count"],
            reverse=True,
        ),
        "models_table": [
            {"name": n, "count": d["count"], "amount": d["amount"]}
            for n, d in models_sorted
        ],
        "model_labels": [n for n, _ in models_sorted[:10]],
        "model_counts": [d["count"] for _, d in models_sorted[:10]],
    }


def _template_ctx(
    request: Request,
    *,
    mode: str,
    days: int,
    month: str | None,
    start_date: date,
    end_date: date,
    period_label: str,
    data: dict,
) -> dict:
    sales_row = data["sales_row"]
    return {
        "request": request,
        "mode": mode,
        "target_date": end_date.isoformat(),
        "days": days,
        "month": month,
        "date_from": start_date.isoformat(),
        "date_to": end_date.isoformat(),
        "period_label": period_label,
        "sales_count": sales_row["count"],
        "preorders_count": data["preorders_row"]["count"],
        "bookings_count": data["bookings_row"]["count"],
        "payment_labels": [
            "Наличные",
            "Терминал",
            "QR",
            "Перевод",
            "По счёту",
            "Рассрочка",
        ],
        "payment_values": [
            float(sales_row["cash"]),
            float(sales_row["terminal"]),
            float(sales_row["qr"]),
            float(sales_row["transfer"]),
            float(sales_row["invoice"]),
            float(sales_row["installment"]),
        ],
        "chart_dates": data["chart_dates"],
        "chart_revenue": data["chart_revenue"],
        "chart_sales": data["chart_sales"],
        "source_labels": data["source_labels"],
        "source_counts": data["source_counts"],
        "source_amounts": data["source_amounts"],
        "sources_table": data["sources_table"],
        "booking_sources_table": data["booking_sources_table"],
        "models_table": data["models_table"],
        "model_labels": data["model_labels"],
        "model_counts": data["model_counts"],
    }


@router.get("/")
async def stats_page(
    request: Request,
    target_date: str | None = None,
    days: int = Query(7, ge=1, le=366),
    mode: str = Query("preset"),
    month: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
):
    start_date, end_date, period_label = _parse_period(
        target_date, days, mode, month, date_from, date_to
    )

    async_session = get_async_session_factory()
    async with async_session() as session:
        data = await _collect_report(session, start_date, end_date)

    return templates.TemplateResponse(
        "stats.html",
        _template_ctx(
            request,
            mode=mode,
            days=days,
            month=month,
            start_date=start_date,
            end_date=end_date,
            period_label=period_label,
            data=data,
        ),
    )


@router.get("/export.xlsx")
async def export_excel(
    target_date: str | None = None,
    days: int = Query(7, ge=1, le=366),
    mode: str = Query("preset"),
    month: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
):
    start_date, end_date, period_label = _parse_period(
        target_date, days, mode, month, date_from, date_to
    )

    async_session = get_async_session_factory()
    async with async_session() as session:
        data = await _collect_report(session, start_date, end_date)

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(color="FFFFFF", bold=True)

    ws = wb.active
    ws.title = "Сводка"
    ws.append(["Период", period_label])
    ws.append([])
    ws.append(["Показатель", "Значение"])
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
    sales_row = data["sales_row"]
    ws.append(["Продажи (шт)", sales_row["count"]])
    ws.append(["Предзаказы (шт)", data["preorders_row"]["count"]])
    ws.append(["Брони (шт)", data["bookings_row"]["count"]])
    ws.append([])
    ws.append(["Оплата", "Сумма ₽"])
    for cell in ws[8]:
        cell.fill = header_fill
        cell.font = header_font
    for label, key in [
        ("Наличные", "cash"),
        ("Терминал", "terminal"),
        ("QR", "qr"),
        ("Перевод", "transfer"),
        ("По счёту", "invoice"),
        ("Рассрочка", "installment"),
    ]:
        ws.append([label, float(sales_row[key])])

    ws2 = wb.create_sheet("Источники")
    ws2.append(["Источник", "Покупок", "Сумма ₽"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in data["sources_table"]:
        ws2.append([row["name"], row["count"], round(row["amount"], 2)])

    ws3 = wb.create_sheet("Топ моделей")
    ws3.append(["Модель / товар", "Кол-во", "Сумма ₽"])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in data["models_table"]:
        ws3.append([row["name"], row["count"], round(row["amount"], 2)])

    ws4 = wb.create_sheet("Брони по площадкам")
    ws4.append(["Площадка", "Броней"])
    for cell in ws4[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in data["booking_sources_table"]:
        ws4.append([row["name"], row["count"]])

    ws5 = wb.create_sheet("По дням")
    ws5.append(["Дата", "Продажи", "Выручка ₽"])
    for cell in ws5[1]:
        cell.fill = header_fill
        cell.font = header_font
    for i, d in enumerate(data["chart_dates"]):
        ws5.append([d, data["chart_sales"][i], data["chart_revenue"][i]])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, min(len(val), 60))
            sheet.column_dimensions[col_letter].width = max(12, max_len + 2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"stats_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/print")
async def stats_print(
    request: Request,
    target_date: str | None = None,
    days: int = Query(7, ge=1, le=366),
    mode: str = Query("preset"),
    month: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
):
    start_date, end_date, period_label = _parse_period(
        target_date, days, mode, month, date_from, date_to
    )
    async_session = get_async_session_factory()
    async with async_session() as session:
        data = await _collect_report(session, start_date, end_date)

    return templates.TemplateResponse(
        "stats_print.html",
        _template_ctx(
            request,
            mode=mode,
            days=days,
            month=month,
            start_date=start_date,
            end_date=end_date,
            period_label=period_label,
            data=data,
        ),
    )
